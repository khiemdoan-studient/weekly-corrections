"""restore_rejection_reasons.py — Restore wiped "Reason for Rejection" values.

Background
----------
Sheet 6 ("Rejected Changes") col O ("Reason for Rejection") was being wiped
on every pipeline run. The cause was `sheets_writer.py::write_corrections`'s
batchClear of `'Rejected Changes'!A:Z`, which included col O.

v2.7.3 narrowed the wipe (A:N) and added hydration from `_RejectedData` col O.
v2.7.4 moved reason storage to a dedicated `_RejectionReasons` tab to also
survive `_migrate_cumulative_tabs`, `_backfill_mismatch_summary`, and
`removeStudentFromCumulativeTabs_` row deletions on Reject toggles.

This script restores reasons from a pre-wipe XLSX export of the corrections
spreadsheet (downloaded by user from File → Version history → Download as
.xlsx) into the v2.7.4 storage tab `_RejectionReasons`.

Schema of `_RejectionReasons`:
    col A = student_id
    col B = reason
    No header row.

Behavior:
    Upsert by student_id. If the tab already has a non-blank reason for a
    student, this script SKIPS that student by default (preserves existing).
    Use --force to overwrite.

Usage
-----
    pip install openpyxl       # one-time, if not already installed
    python restore_rejection_reasons.py path/to/pre-wipe-export.xlsx
    python restore_rejection_reasons.py path/to/pre-wipe-export.xlsx --force

Output
------
    Restored: N reasons written to _RejectionReasons
    Preserved: K (existing non-blank reason — pass --force to overwrite)
"""

import argparse
import sys

from google.oauth2 import service_account
from googleapiclient.discovery import build

from config import (
    SERVICE_ACCOUNT_KEY,
    SCOPES,
    OUTPUT_SPREADSHEET_ID,
)
from retry_helper import retry_api

# Sheet 6 (XLSX-side) layout — 1-based for openpyxl:
#   col M (13) = Student_ID
#   col O (15) = Reason for Rejection
#   data rows start at row 7 (1-6 are title/caption/filter/sort/header/spacer)
SHEET6_TAB = "Rejected Changes"
SHEET6_STUDENT_ID_COL = 13
SHEET6_REASON_COL = 15
SHEET6_DATA_START_ROW = 7

# `_RejectionReasons` (Sheets-side) layout — 1-based:
#   col A (1) = student_id
#   col B (2) = reason
REASONS_TAB = "_RejectionReasons"
REASONS_SID_COL_1B = 1
REASONS_REASON_COL_1B = 2


def _read_pre_wipe_xlsx(xlsx_path):
    """Parse the user-downloaded XLSX and return {student_id: reason}.

    Skips empty reasons. Strips whitespace. Last-write-wins on duplicates.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print(
            "  ERROR: openpyxl not installed. Run: pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    if SHEET6_TAB not in wb.sheetnames:
        print(
            f"  ERROR: workbook does not contain a '{SHEET6_TAB}' tab.\n"
            f"  Available tabs: {wb.sheetnames}",
            file=sys.stderr,
        )
        sys.exit(1)

    ws = wb[SHEET6_TAB]
    reasons = {}
    for row in ws.iter_rows(
        min_row=SHEET6_DATA_START_ROW,
        min_col=1,
        max_col=SHEET6_REASON_COL,
        values_only=True,
    ):
        if len(row) < SHEET6_REASON_COL:
            continue
        sid = str(row[SHEET6_STUDENT_ID_COL - 1] or "").strip()
        reason = str(row[SHEET6_REASON_COL - 1] or "").strip()
        if not sid or not reason:
            continue
        reasons[sid] = reason

    wb.close()
    return reasons


def _ensure_reasons_tab(sheets_service):
    """Ensure `_RejectionReasons` tab exists. Hide it on creation. Returns sheetId."""
    resp = retry_api(
        lambda: sheets_service.spreadsheets()
        .get(spreadsheetId=OUTPUT_SPREADSHEET_ID, fields="sheets.properties")
        .execute(),
        label="get spreadsheet metadata",
    )
    for sheet in resp.get("sheets", []):
        props = sheet["properties"]
        if props["title"] == REASONS_TAB:
            return props["sheetId"]

    # Create + hide
    print(f"  Creating tab '{REASONS_TAB}' (hidden)...")
    create_resp = retry_api(
        lambda: sheets_service.spreadsheets()
        .batchUpdate(
            spreadsheetId=OUTPUT_SPREADSHEET_ID,
            body={
                "requests": [
                    {"addSheet": {"properties": {"title": REASONS_TAB, "hidden": True}}}
                ]
            },
        )
        .execute(),
        label="create _RejectionReasons",
    )
    return create_resp["replies"][0]["addSheet"]["properties"]["sheetId"]


def _read_existing_reasons(sheets_service):
    """Returns dict {student_id: (1-based row_num, existing_reason)}."""
    resp = retry_api(
        lambda: sheets_service.spreadsheets()
        .values()
        .get(
            spreadsheetId=OUTPUT_SPREADSHEET_ID,
            range=f"'{REASONS_TAB}'!A:B",
        )
        .execute(),
        label="read _RejectionReasons",
    )
    rows = resp.get("values", [])
    out = {}
    for i, row in enumerate(rows):
        sid = str(row[0] or "").strip() if len(row) >= 1 else ""
        existing = str(row[1] or "").strip() if len(row) >= 2 else ""
        if sid:
            out[sid] = (i + 1, existing)  # 1-based row number
    return out


def _apply_writes(sheets_service, updates, appends):
    """updates: list of (row_num, reason). appends: list of (sid, reason)."""
    if updates:
        data = [
            {"range": f"'{REASONS_TAB}'!B{row_num}", "values": [[reason]]}
            for row_num, reason in updates
        ]
        retry_api(
            lambda: sheets_service.spreadsheets()
            .values()
            .batchUpdate(
                spreadsheetId=OUTPUT_SPREADSHEET_ID,
                body={"valueInputOption": "RAW", "data": data},
            )
            .execute(),
            label="update _RejectionReasons col B",
        )

    if appends:
        retry_api(
            lambda: sheets_service.spreadsheets()
            .values()
            .append(
                spreadsheetId=OUTPUT_SPREADSHEET_ID,
                range=f"'{REASONS_TAB}'!A:B",
                valueInputOption="RAW",
                insertDataOption="INSERT_ROWS",
                body={"values": [[sid, reason] for sid, reason in appends]},
            )
            .execute(),
            label="append to _RejectionReasons",
        )


def main():
    parser = argparse.ArgumentParser(
        description="Restore wiped Reason values to _RejectionReasons from a pre-wipe XLSX export.",
    )
    parser.add_argument(
        "xlsx_path",
        help="Path to the pre-wipe XLSX export (from Sheets Version History).",
    )
    parser.add_argument(
        "--force",
        action="store_true",
        help="Overwrite existing non-blank reasons in _RejectionReasons. Default: skip them.",
    )
    args = parser.parse_args()

    print("=" * 60)
    print("  RESTORE — Reason for Rejection from pre-wipe XLSX (v2.7.4)")
    print("=" * 60)

    print(f"\n  Reading XLSX: {args.xlsx_path}")
    pre_wipe = _read_pre_wipe_xlsx(args.xlsx_path)
    print(f"  -> {len(pre_wipe)} (student_id, reason) pair(s) found in pre-wipe export")

    if not pre_wipe:
        print("\n  No reasons to restore. Exiting.")
        return

    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_KEY, scopes=SCOPES
    )
    sheets = build("sheets", "v4", credentials=creds)

    _ensure_reasons_tab(sheets)

    print(f"\n  Reading current {REASONS_TAB}...")
    existing = _read_existing_reasons(sheets)
    print(f"  -> {len(existing)} row(s) currently in {REASONS_TAB}")

    updates = []  # (row_num, reason) — for sids already in the tab
    appends = []  # (sid, reason) — for sids not yet in the tab
    preserved = []  # skipped because existing non-blank reason

    for sid, reason in pre_wipe.items():
        match = existing.get(sid)
        if match is None:
            appends.append((sid, reason))
        else:
            row_num, existing_reason = match
            if existing_reason and not args.force:
                preserved.append((sid, existing_reason, reason))
            else:
                updates.append((row_num, reason))

    print(f"\n  Plan:")
    print(f"    Append (new sid)              : {len(appends)}")
    print(f"    Update (existing sid, blank)  : {len(updates)}")
    print(f"    Preserved (existing non-blank): {len(preserved)}")

    if not updates and not appends:
        print("\n  Nothing to write. Exiting without changes.")
        return

    print(f"\n  Writing to {REASONS_TAB}...")
    _apply_writes(sheets, updates, appends)
    print(f"  Done. {len(updates) + len(appends)} reason(s) written.")

    if preserved:
        print(
            f"\n  {len(preserved)} reason(s) preserved (existing non-blank value). "
            f"Pass --force to overwrite."
        )

    print(
        f"\n  Next pipeline run will hydrate Sheet 6 col O from "
        f"{REASONS_TAB}. To test now: python generate_corrections.py"
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n  FATAL: {e}")
        sys.exit(1)
